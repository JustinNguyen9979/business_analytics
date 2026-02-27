import json, crud, models, schemas, standard_parser, secrets, string
from services.search_service import search_service
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Response, status, Query, Body, Request
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.responses import ORJSONResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from typing import List, Dict, Any, Union
from database import SessionLocal, engine
from datetime import date, timedelta
import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Alignment
from cache import redis_client
from celery_worker import process_data_request, recalculate_all_brand_data, recalculate_brand_data_specific_dates
from slowapi import _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from slowapi.middleware import SlowAPIMiddleware
from limiter import limiter
from auth_utils import get_password_hash, verify_password, create_access_token, SECRET_KEY, ALGORITHM
from fastapi.security import OAuth2PasswordBearer
from jose import JWTError, jwt

app = FastAPI(
    # title="CEO Dashboard API by Julice",
    default_response_class=ORJSONResponse 
)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="auth/login")

app.state.limiter = limiter
app.add_middleware(SlowAPIMiddleware)
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# models.Base.metadata.create_all(bind=engine)

app.add_middleware(GZipMiddleware, minimum_size=1000)

# --- Dependencies ---
def get_db(): 
    db = SessionLocal()
    try: 
        yield db 
    finally: 
        db.close()

async def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Không thể xác thực thông tin đăng nhập",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
        token_data = schemas.TokenData(username=username, role=payload.get("role"))
    except JWTError:
        raise credentials_exception
    
    user = db.query(models.User).filter(models.User.username == token_data.username).first()
    if user is None:
        raise credentials_exception
    return user

class RoleChecker:
    def __init__(self, allowed_roles: List[models.UserRole]):
        self.allowed_roles = allowed_roles

    def __call__(self, user: models.User = Depends(get_current_user)):
        if user.role not in self.allowed_roles:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Bạn không có quyền thực hiện hành động này"
            )
        return user

# ==============================================================================
# === 0. AUTHENTICATION ENDPOINTS (SIGNUP / LOGIN) ===
# ==============================================================================

@app.post("/api/auth/signup", response_model=schemas.User, status_code=status.HTTP_201_CREATED)
def signup(user_in: schemas.UserCreate, db: Session = Depends(get_db)):
    """Đăng ký tài khoản mới - Luôn gán role ADMIN và tạo ID ngẫu nhiên"""
    # Kiểm tra username tồn tại
    user = db.query(models.User).filter(models.User.username == user_in.username).first()
    if user:
        raise HTTPException(status_code=400, detail="Username đã tồn tại.")
    
    # Kiểm tra email tồn tại
    user = db.query(models.User).filter(models.User.email == user_in.email).first()
    if user:
        raise HTTPException(status_code=400, detail="Email đã được sử dụng.")
    
    # Tạo mã định danh ngẫu nhiên (12 ký tự)
    alphabet = string.ascii_letters + string.digits
    random_id = ''.join(secrets.choice(alphabet) for _ in range(12))
    
    # Tạo user mới với role ADMIN
    hashed_password = get_password_hash(user_in.password)
    db_user = models.User(
        id=random_id, # Sử dụng ID ngẫu nhiên
        username=user_in.username,
        email=user_in.email,
        full_name=user_in.full_name,
        hashed_password=hashed_password,
        role=models.UserRole.ADMIN 
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user

@app.post("/api/auth/login", response_model=schemas.Token)
def login(form_data: schemas.UserLogin, db: Session = Depends(get_db)): 
    """Đăng nhập và nhận Token"""
    user = db.query(models.User).filter(models.User.username == form_data.username).first()
    if not user or not verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Tài khoản hoặc mật khẩu không chính xác.",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token = create_access_token(data={"sub": user.username, "role": user.role.value})
    return {"access_token": access_token, "token_type": "bearer"}

def get_brand_from_slug(brand_slug: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    """Lấy brand từ slug và kiểm tra quyền sở hữu"""
    # [TỐI ƯU] Tìm thẳng theo slug + owner_id để đảm bảo lấy đúng brand của user
    db_brand = crud.get_brand_by_slug(db, slug=brand_slug, owner_id=current_user.id)
    if not db_brand:
        raise HTTPException(status_code=404, detail="Không tìm thấy Brand.")
    
    return db_brand

# ==============================================================================
# === 1. ENDPOINTS QUẢN LÝ THƯƠNG HIỆU (BRAND MANAGEMENT) ===
# ==============================================================================

@app.get("/api/", response_model=schemas.MessageResponse)
def read_root(): 
    return {"message": "Chào mừng đến với CEO Dashboard API!"}

@app.get("/api/brands/", response_model=List[schemas.BrandInfo])
def read_brands(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)): 
    """Lấy danh sách các brand thuộc sở hữu của user hiện tại"""
    return db.query(models.Brand).filter(models.Brand.owner_id == current_user.id).all()

@app.post("/api/brands/", response_model=schemas.BrandInfo)
def create_brand_api(brand: schemas.BrandCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    """Tạo brand mới gắn với user hiện tại"""
    # [TỐI ƯU] Kiểm tra tên trùng trong phạm vi sở hữu của user
    existing = db.query(models.Brand).filter(
        models.Brand.name == brand.name,
        models.Brand.owner_id == current_user.id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Bạn đã có một Brand với tên này rồi.")
        
    new_brand = crud.create_brand(db=db, obj_in=brand, owner_id=current_user.id)
    return new_brand

@app.put("/api/brands/{brand_id}", response_model=schemas.BrandInfo)
def update_brand_api(brand_id: int, brand_update: schemas.BrandCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    # Kiểm tra quyền sở hữu
    db_brand = db.query(models.Brand).filter(models.Brand.id == brand_id, models.Brand.owner_id == current_user.id).first()
    if not db_brand:
        raise HTTPException(status_code=403, detail="Bạn không có quyền chỉnh sửa thương hiệu này.")
        
    updated_brand = crud.update_brand_name(db, brand_id=brand_id, new_name=brand_update.name, owner_id=current_user.id)
    if not updated_brand:
        raise HTTPException(status_code=400, detail="Không thể đổi tên. Tên Brand mới đã bị trùng trong danh sách của bạn.")
    return updated_brand

@app.delete("/api/brands/{brand_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_brand_api(brand_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    # Kiểm tra quyền sở hữu
    db_brand = db.query(models.Brand).filter(models.Brand.id == brand_id, models.Brand.owner_id == current_user.id).first()
    if not db_brand:
        raise HTTPException(status_code=403, detail="Bạn không có quyền xóa thương hiệu này.")
        
    deleted_brand = crud.delete_brand_by_id(db, brand_id=brand_id)
    if not deleted_brand:
        raise HTTPException(status_code=404, detail="Không tìm thấy Brand để xóa.")
    return Response(status_code=status.HTTP_204_NO_CONTENT)

@app.post("/api/brands/{brand_id}/clone", response_model=schemas.BrandInfo)
def clone_brand_api(brand_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    # Kiểm tra quyền sở hữu
    db_brand = db.query(models.Brand).filter(models.Brand.id == brand_id, models.Brand.owner_id == current_user.id).first()
    if not db_brand:
        raise HTTPException(status_code=403, detail="Bạn không có quyền nhân bản thương hiệu này.")
        
    cloned = crud.clone_brand(db, brand_id=brand_id)
    if not cloned:
        raise HTTPException(status_code=404, detail="Không tìm thấy Brand để nhân bản.")
    return cloned

@app.post("/api/brands/{brand_slug}/trigger-recalculation", status_code=status.HTTP_202_ACCEPTED, response_model=schemas.MessageResponse)
@limiter.limit("5/minute")
def trigger_recalculation_api(request: Request, brand: models.Brand = Depends(get_brand_from_slug)):
    """
    Kích hoạt task chạy nền để tính toán lại toàn bộ dữ liệu cho brand.
    Trả về ngay lập tức.
    """
    recalculate_all_brand_data.delay(brand.id)
    return {"message": "Yêu cầu tính toán lại đã được gửi đi. Dữ liệu sẽ được cập nhật trong nền."}


# ==============================================================================
# === 2. ENDPOINTS XỬ LÝ DỮ LIỆU (DATA PROCESSING) ===
# ==============================================================================

@app.post("/api/brands/{brand_slug}/upload-standard-file", status_code=status.HTTP_202_ACCEPTED, response_model=schemas.MessageResponse)
@limiter.limit("5/minute")
async def upload_standard_file(
    request: Request,
    platform: str, 
    brand: models.Brand = Depends(get_brand_from_slug),
    db: Session = Depends(get_db), 
    file: UploadFile = File(...),
    force: bool = Query(False, description="Nếu True, sẽ xử lý lại file ngay cả khi đã tồn tại trong lịch sử import.")
):
    """Nhận file, xử lý và kích hoạt worker tính toán lại toàn bộ."""
    result = standard_parser.process_standard_file(
        db, 
        await file.read(), 
        brand.id, 
        platform, 
        file_name=file.filename,
        allow_override=force
    )
    if result.get("status") == "error":
        raise HTTPException(status_code=400, detail=result.get("message"))
    
    # [TỐI ƯU] Kích hoạt task tính toán lại thông minh
    affected_dates = result.get("affected_dates")
    task = None
    
    if affected_dates and isinstance(affected_dates, list) and len(affected_dates) > 0:
        print(f"API: Kích hoạt tính toán lại cho {len(affected_dates)} ngày cụ thể.")
        task = recalculate_brand_data_specific_dates.delay(brand.id, affected_dates)
    else:
        # Fallback: Nếu không xác định được ngày (file rỗng?), tính lại toàn bộ cho chắc
        print("API: Không xác định được ngày cụ thể, tính toán lại TOÀN BỘ.")
        task = recalculate_all_brand_data.delay(brand.id)
    
    # [UX IMPROVEMENT] Chờ worker hoàn thành (tối đa 60s) để Frontend hiển thị Loading đúng thực tế
    if task:
        try:
            # Chờ task hoàn thành. Vì đã tối ưu nên thường chỉ mất < 5s.
            task.get(timeout=60)
            print("API: Worker đã hoàn thành nhiệm vụ tính toán.")
        except Exception as e:
            print(f"API: Worker chưa kịp hoàn thành trong 60s hoặc có lỗi: {e}. Client sẽ tự refresh.")
            # Không raise error để tránh làm user hoang mang, worker vẫn sẽ chạy ngầm tiếp.
        
    return {"message": "Upload và xử lý dữ liệu thành công!"}

@app.post("/api/brands/{brand_slug}/recalculate-and-wait", status_code=status.HTTP_200_OK, response_model=schemas.MessageResponse)
@limiter.limit("3/minute")
def recalculate_and_wait(request: Request, brand: models.Brand = Depends(get_brand_from_slug)):
    """
    Kích hoạt Worker để tính toán lại và BLOCK cho đến khi task hoàn thành.
    """
    print(f"API: Nhận yêu cầu recalculate-and-wait cho brand {brand.id}.")
    
    # Gửi task đến Worker và lấy về đối tượng AsyncResult
    task_result = recalculate_all_brand_data.delay(brand.id)
    
    try:
        # Dòng quan trọng: Chờ đợi kết quả của task, với timeout là 5 phút (300 giây)
        task_result.get(timeout=300) 
        print(f"API: Worker đã hoàn thành recalculate cho brand {brand.id}.")
        return {"message": "Tính toán lại hoàn tất!"}
    except TimeoutError:
        raise HTTPException(status_code=504, detail="Quá trình tính toán mất quá nhiều thời gian.")

class DateRangePayload(BaseModel):
    start_date: date
    end_date: date

@app.get("/api/brands/{brand_slug}/sources", response_model=List[str])
def get_brand_sources(brand: models.Brand = Depends(get_brand_from_slug), db: Session = Depends(get_db)):
    """Lấy danh sách tất cả các 'source' duy nhất cho một brand."""
    return crud.get_sources_for_brand(db, brand.id)

@app.post("/api/brands/{brand_slug}/delete-data-in-range", status_code=status.HTTP_200_OK, response_model=schemas.DeleteDataResponse)
@app.post("/api/brands/{brand_slug}/delete-data", response_model=schemas.DeleteDataResponse)
@limiter.limit("5/minute")
def delete_data_in_range(
    request: Request,
    payload: DateRangePayload,
    brand: models.Brand = Depends(get_brand_from_slug),
    db: Session = Depends(get_db),
    source: str = Query(None, description="Optional: Filter by source (e.g., 'shopee', 'tiktok'). If not provided, deletes from all sources.")
):
    """
    Deletes transactional data for a brand within a specified date range.
    Returns a list of sources that have been completely wiped out.
    """
    try:
        # Hàm crud giờ sẽ trả về danh sách các source đã bị xóa hoàn toàn
        fully_deleted_sources = crud.delete_brand_data_in_range(
            db, brand.id, payload.start_date, payload.end_date, source
        )
        
        # Đảm bảo trả về mảng rỗng nếu fully_deleted_sources là None
        if fully_deleted_sources is None:
            fully_deleted_sources = []
            
        # Kích hoạt tính toán lại dữ liệu trong nền (Chỉ tính các ngày bị ảnh hưởng)
        from datetime import timedelta
        
        delta = payload.end_date - payload.start_date
        affected_dates = []
        # Tạo danh sách các ngày trong khoảng thời gian xóa
        for i in range(delta.days + 1):
            day = payload.start_date + timedelta(days=i)
            affected_dates.append(day.isoformat())
            
        if affected_dates:
            recalculate_brand_data_specific_dates.delay(brand.id, affected_dates)
        else:
            # Fallback nếu không xác định được ngày
            recalculate_all_brand_data.delay(brand.id)
        
        return {
            "message": "Yêu cầu xóa dữ liệu đã được thực hiện. Dữ liệu đang được tính toán lại.",
            "fully_deleted_sources": fully_deleted_sources
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi server khi xóa dữ liệu: {str(e)}")

def generate_cache_key(brand_id: int, request_type: str, params: Dict[str, Any]) -> str:
    """Tạo ra một cache key nhất quán từ thông tin request."""
    param_string = ":".join(f"{k}={v}" for k, v in sorted(params.items()))
    return f"data_req:{brand_id}:{request_type}:{param_string}"

@app.post("/api/data-requests", status_code=status.HTTP_202_ACCEPTED, response_model=schemas.TaskResponse)
@limiter.limit("60/minute")
def request_data_processing(
    request: Request,
    request_body: schemas.DataRequest, # Sử dụng Pydantic model để xác thực payload
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Endpoint chính: Nhận yêu cầu (dùng SLUG), kiểm tra cache, hoặc giao việc cho worker.
    """
    # 1. Lấy brand và kiểm tra quyền sở hữu (TỐI ƯU)
    db_brand = crud.get_brand_by_slug(db, slug=request_body.brand_slug, owner_id=current_user.id)
    if not db_brand:
        raise HTTPException(status_code=404, detail=f"Brand slug '{request_body.brand_slug}' không tồn tại hoặc bạn không có quyền truy cập.")
    
    brand_id = db_brand.id
    request_type = request_body.request_type
    params = request_body.params

    cache_key = generate_cache_key(brand_id, request_type, params)
    
    # Bước 1: Kiểm tra cache
    cached_result = redis_client.get(cache_key)
    if cached_result:
        print(f"API: Cache HIT cho key: {cache_key}")
        # Trả về ngay lập tức nếu tìm thấy
        return ORJSONResponse(
            content={"status": "SUCCESS", "data": json.loads(cached_result)},
            status_code=status.HTTP_200_OK
        )
    
    # Bước 2: Cache miss -> Giao việc cho worker
    print(f"API: Cache MISS cho key: {cache_key}. Giao việc cho worker.")
    task = process_data_request.delay(
        request_type=request_type,
        cache_key=cache_key,
        brand_id=brand_id,
        params=params
    )
    
    # Trả về task_id để frontend có thể "hỏi thăm"
    return {"task_id": task.id, "status": "PROCESSING", "cache_key": cache_key}

@app.get("/api/data-requests/status/{cache_key}", response_model=schemas.TaskStatusResponse)
def get_request_status(
    cache_key: str, 
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Endpoint để Frontend "hỏi thăm" xem dữ liệu đã được xử lý xong chưa.
    Nó kiểm tra cache và quyền truy cập dựa trên brand_id trong cache_key.
    """
    # Parse brand_id từ cache_key (định dạng: data_req:brand_id:request_type:params)
    try:
        parts = cache_key.split(":")
        if len(parts) < 3:
            raise ValueError("Invalid cache key format")
        brand_id = int(parts[1])
    except (ValueError, IndexError):
        raise HTTPException(status_code=400, detail="Cache key không hợp lệ.")

    # Kiểm tra quyền sở hữu brand
    db_brand = db.query(models.Brand).filter(models.Brand.id == brand_id).first()
    if not db_brand or db_brand.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Bạn không có quyền xem dữ liệu này.")

    cached_result = redis_client.get(cache_key)
    
    if cached_result:
        print(f"API: Polling HIT cho key: {cache_key}")
        result = json.loads(cached_result)
        # Kiểm tra xem worker có báo lỗi không
        if isinstance(result, dict) and result.get("status") == "FAILED":
            return {"status": "FAILED", "error": result.get("error", "Lỗi không xác định từ worker.")}
        # Thành công
        return {"status": "SUCCESS", "data": result}
    else:
        # Vẫn đang xử lý
        return {"status": "PROCESSING"}

@app.post("/api/brands/{brand_slug}/recalculate", status_code=status.HTTP_200_OK, response_model=schemas.RecalculationResponse)
@limiter.limit("5/minute")
def recalculate_brand_data(request: Request, brand: models.Brand = Depends(get_brand_from_slug), db: Session = Depends(get_db)):
    """Tính toán lại toàn bộ dữ liệu của một brand một cách đồng bộ."""
    result = crud.recalculate_brand_data_sync(db, brand_id=brand.id)
    return result

@app.get("/api/brands/{brand_slug}/download-sample-file")
def download_sample_file(brand: models.Brand = Depends(get_brand_from_slug)):
    """Tạo và trả về file Excel mẫu với định dạng chuyên nghiệp: màu sắc và kích thước cột."""
    output = io.BytesIO()
    
    # Định nghĩa style
    note_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Xanh nhạt
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Xanh đậm
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 1. Sheet Giá vốn
        df_cost = pd.DataFrame([
            {'sku': 'SP001', 'name': 'Sản phẩm mẫu A', 'cost_price': 100000},
            {'sku': 'SP002', 'name': 'Sản phẩm mẫu B', 'cost_price': 150000}
        ])
        df_cost.to_excel(writer, sheet_name='Giá vốn', index=False, startrow=1)
        ws = writer.sheets['Giá vốn']
        notes = ["Mã SKU (Bắt buộc)", "Tên Sản Phẩm", "Giá Vốn Nhập Kho"]
        
        # Thiết lập độ rộng cột
        widths = {'A': 30, 'B': 60, 'C': 20}
        for col, width in widths.items(): ws.column_dimensions[col].width = width
        
        # Style cho Row 1 (Notes) và Row 2 (Headers)
        for col_num, note in enumerate(notes, 1):
            cell_n = ws.cell(row=1, column=col_num)
            cell_n.value = note
            cell_n.fill = note_fill
            cell_n.font = bold_font
            cell_n.alignment = center_align
            
            cell_h = ws.cell(row=2, column=col_num)
            cell_h.fill = header_fill
            cell_h.font = white_font
            cell_h.alignment = center_align

        # 2. Sheet Đơn hàng
        df_order = pd.DataFrame([{
            'order_id': 'ORD12345', 
            'order_status': 'Completed', 
            'cancel_reason': '',
            'tracking_id': 'ORD12345',
            'order_date': '2024-01-01 10:00:00', 
            'shipped_time': '2024-01-02 14:00:00', 
            'delivered_date': '2024-01-03 15:00:00',
            'sku': 'SP001', 
            'quantity': 2, 
            'original_price': 250000, 
            'subsidy_amount': 10000,
            'sku_price': 200000, 
            'username': 'customer_01', 
            'payment_method': 'COD', 
            'shipping_provider_name': 'Shopee Xpress', 'province': 'Hà Nội', 'district': 'Quận Cầu Giấy', 'address': 'Số 1 Duy Tân',
            'phone': '0912345678', 'email': 'khachhang@example.com', 'gender': 'Nữ'
        }])
        df_order.to_excel(writer, sheet_name='Đơn hàng', index=False, startrow=1)
        ws = writer.sheets['Đơn hàng']
        notes = ["Mã Đơn Hàng", "Trạng Thái", "Lý Do Hủy", "Vận Đơn", "Ngày Đặt Hàng",  "Ngày Gửi Kho", "Ngày Giao", "SKU", "Số Lượng", "Giá Niêm Yết", "Trợ Giá", "Giá Bán", "Khách Hàng", "Thanh Toán", "Vận Chuyển", "Tỉnh", "Quận", "Địa Chỉ", "SĐT", "Email", "Giới Tính"]
        
        # Thiết lập độ rộng cột (A-U)
        for i, width in enumerate([20, 22, 25, 22, 20, 20, 20, 20, 10, 15, 15, 15, 15, 20, 20, 15, 15, 15, 25, 30, 10], 1):
            ws.column_dimensions[chr(64+i) if i <= 26 else 'A'+chr(64+i-26)].width = width

        for col_num, note in enumerate(notes, 1):
            ws.cell(row=1, column=col_num).value = note
            ws.cell(row=1, column=col_num).fill = note_fill
            ws.cell(row=1, column=col_num).font = bold_font
            ws.cell(row=2, column=col_num).fill = header_fill
            ws.cell(row=2, column=col_num).font = white_font

        # 3. Sheet Doanh thu
        df_revenue = pd.DataFrame([{
            'order_id': 'ORD12345', 'order_refund': 'ORD-RE12345', 'order_date': '2024-01-01', 'transaction_date': '2024-01-04',
            'net_revenue': 390000, 'gmv': 500000, 'total_fees': 50000, 'refund': 0
        }])
        df_revenue.to_excel(writer, sheet_name='Doanh thu', index=False, startrow=1)
        ws = writer.sheets['Doanh thu']
        notes = ["Mã Đơn Hàng", "Mã Đơn Hoàn", "Ngày Đặt", "Ngày Tiền Về", "Thực Nhận", "GMV", "Phí Sàn", "Hoàn Trả"]
        for i, width in enumerate([25, 30, 18, 18, 15, 15, 15, 15], 1): ws.column_dimensions[chr(64+i)].width = width
        for col_num, note in enumerate(notes, 1):
            ws.cell(row=1, column=col_num).value = note
            ws.cell(row=1, column=col_num).fill = note_fill
            ws.cell(row=1, column=col_num).font = bold_font
            ws.cell(row=2, column=col_num).fill = header_fill
            ws.cell(row=2, column=col_num).font = white_font

        # 4. Sheet Marketing
        df_marketing = pd.DataFrame([{
            'ads_date': '2024-01-01', 'adSpend': 500000, 'conversion': 10, 'impressions': 1000, 'click': 100
        }])
        df_marketing.to_excel(writer, sheet_name='Marketing', index=False, startrow=1)
        ws = writer.sheets['Marketing']
        notes = ["Ngày Chạy QC", "Chi Phí", "Chuyển Đổi", "Hiển Thị", "Click"]
        for i, width in enumerate([18, 15, 12, 12, 12], 1): ws.column_dimensions[chr(64+i)].width = width
        for col_num, note in enumerate(notes, 1):
            ws.cell(row=1, column=col_num).value = note
            ws.cell(row=1, column=col_num).fill = note_fill
            ws.cell(row=1, column=col_num).font = bold_font
            ws.cell(row=2, column=col_num).fill = header_fill
            ws.cell(row=2, column=col_num).font = white_font

    output.seek(0)
    filename = f"Template_Bao_Cao_{brand.slug}.xlsx"
    return StreamingResponse(
        output, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

# ==============================================================================
# === 3. ENDPOINTS LẤY DỮ LIỆU CHO DASHBOARD (DATA RETRIEVAL) ===
# ==============================================================================

@app.get("/api/brands/{brand_slug}/customer-map-distribution", response_model=List[schemas.CustomerMapDistributionItem])
def read_customer_map_distribution(
    start_date: date,
    end_date: date,
    status: List[str] = Query(['completed'], description="Trạng thái đơn hàng cần lấy (completed, cancelled, bomb, refunded). Mặc định chỉ lấy completed."),
    source: List[str] = Query(None, description="Danh sách nguồn dữ liệu (e.g. shopee, lazada)"),
    brand: models.Brand = Depends(get_brand_from_slug),
    db: Session = Depends(get_db)
):
    """Lấy dữ liệu phân bổ khách hàng theo tỉnh/thành để vẽ bản đồ."""
    try:
        distribution_data = crud.get_aggregated_location_distribution(
            db, brand.id, start_date, end_date, status_filter=status, source_list=source
        )
        return distribution_data
    except Exception as e:
        print(f"!!! LỖI ENDPOINT CUSTOMER MAP: {e}")
        # Trả về list rỗng thay vì lỗi 500 để tránh crash UI
        return []

@app.get("/api/brands/{brand_slug}", response_model=schemas.BrandWithKpis)
def read_brand_kpis(
    start_date: date, 
    end_date: date, 
    brand: models.Brand = Depends(get_brand_from_slug),
    db: Session = Depends(get_db)
):
    """Lấy thông tin tổng quan và các chỉ số KPI tổng hợp của một brand."""
    db_brand, cache_was_missing = crud.get_brand_details(db, brand.id, start_date, end_date)
    if not db_brand: 
        raise HTTPException(status_code=404, detail="Không tìm thấy Brand.")
    if cache_was_missing:
        print(f"API: Phát hiện cache lỗi thời cho brand ID {brand.id}. Dashboard có thể hiển thị dữ liệu cũ.")
    
    # [TỐI ƯU] Đã loại bỏ lệnh recalculate_all_brand_data vô điều kiện.
    # Chỉ tính lại khi có hành động cụ thể (Upload/Delete).
    return db_brand

@app.get("/api/brands/{brand_slug}/daily-kpis", response_model=schemas.DailyKpiResponse)
def read_brand_daily_kpis(
    start_date: date, 
    end_date: date, 
    source: List[str] = Query(None), # Thêm tham số source
    brand: models.Brand = Depends(get_brand_from_slug),
    db: Session = Depends(get_db)
):
    """Lấy dữ liệu KPI hàng ngày cho việc vẽ biểu đồ, có hỗ trợ lọc theo nguồn."""
    daily_data = crud.get_daily_kpis_for_range(db, brand.id, start_date, end_date, source_list=source)
    return {"data": daily_data}

@app.get("/api/brands/{brand_slug}/top-products", response_model=List[schemas.TopProduct])
def read_top_products(
    start_date: date,
    end_date: date,
    source: List[str] = Query(None, description="Lọc theo nguồn (shopee, lazada...)"),
    brand: models.Brand = Depends(get_brand_from_slug),
    limit: int = 10,
    db: Session = Depends(get_db)
):
    """Lấy top N sản phẩm bán chạy nhất."""
    try:
        top_products = crud.get_top_selling_products(db, brand.id, start_date, end_date, limit, source_list=source)
        return top_products
    except Exception as e:
        print(f"!!! LỖI ENDPOINT TOP PRODUCTS: {e}")
        raise HTTPException(status_code=500, detail="Lỗi server khi xử lý yêu cầu.")


@app.get("/api/brands/{brand_slug}/kpis/operation", response_model=schemas.OperationKpisResponse)
@limiter.limit("30/minute")
def get_operation_kpis (
    request: Request,
    brand_slug: str,
    start_date: date = Query(..., description="Ngày bắt đầu (YYYY-MM-DD)"),
    end_date: date = Query(..., description="Ngày kết thúc (YYYY-MM-DD)"),
    source: List[str] = Query(None, description="Danh sách nguồn dữ liệu (e.g. shopee, lazada)"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
): 
    """
    Lấy các chỉ số KPI vận hành tổng hợp cho OperationPage.
    Trả về giá trị trung bình trong khoảng thời gian được chọn.
    """
    db_brand = crud.get_brand_by_slug(db, slug=brand_slug, owner_id=current_user.id)
    if not db_brand:
        raise HTTPException(status_code=404, detail="Không tìm thấy Brand.")
    
    # Lấy dữ liệu KPI theo ngày trong range
    # Nếu không có source truyền vào, mặc định là ['all'] bên trong logic
    
    print(f"DEBUG_API: get_operation_kpis called with range: {start_date} -> {end_date}, sources: {source}")
    
    # Sử dụng hàm aggregation mới từ crud (đã dùng kpi_utils để gộp JSON)
    operation_kpis = crud.get_aggregated_operation_kpis(
        db, 
        db_brand.id, 
        start_date, 
        end_date, 
        source_list=source
    )

    # Convert kết quả dict thành Pydantic model response
    return operation_kpis

@app.get("/api/brands/{brand_slug}/kpis/customer", response_model=schemas.CustomerKpisResponse)
@limiter.limit("30/minute")
def get_customer_kpis (
    request: Request,
    brand_slug: str,
    start_date: date = Query(..., description="Ngày bắt đầu (YYYY-MM-DD)"),
    end_date: date = Query(..., description="Ngày kết thúc (YYYY-MM-DD)"),
    source: List[str] = Query(None, description="Danh sách nguồn dữ liệu"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Lấy các chỉ số KPI khách hàng cho CustomerPage.
    """
    db_brand = crud.get_brand_by_slug(db, slug=brand_slug, owner_id=current_user.id)
    if not db_brand:
        raise HTTPException(status_code=404, detail="Không tìm thấy Brand.")
        
    customer_kpis = crud.get_aggregated_customer_kpis(
        db,
        db_brand.id,
        start_date,
        end_date,
        source_list=source
    )
    return customer_kpis

@app.get("/api/brands/{brand_slug}/customers", response_model=List[schemas.CustomerAnalyticsItem])
@limiter.limit("30/minute")
def get_top_customers(
    request: Request,
    brand_slug: str,
    limit: int = 50,
    sort_by: str = Query("total_spent", description="Trường cần sắp xếp: total_spent, total_orders, bomb_orders..."),
    order: str = Query("desc", description="Thứ tự: asc hoặc desc"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Lấy danh sách Top Khách hàng (Đã chuyển sang tính toán động).
    Mặc định lấy toàn bộ lịch sử (2020 -> 2030) nếu không có bộ lọc ngày.
    """
    db_brand = crud.get_brand_by_slug(db, slug=brand_slug, owner_id=current_user.id)
    if not db_brand:
        raise HTTPException(status_code=404, detail="Không tìm thấy Brand.")
    
    # Giả lập khoảng thời gian "Toàn thời gian"
    start_date = date(2020, 1, 1)
    end_date = date(2030, 12, 31)
    
    # Gọi hàm tính toán động
    # Lưu ý: Hàm này trả về CustomerPaginationResponse, ta chỉ cần list data
    result = crud.customer.get_top_customers_in_period(
        db, 
        brand_id=db_brand.id, 
        start_date=start_date, 
        end_date=end_date, 
        limit=limit
    )
    
    # Kết quả trả về là List[CustomerAnalyticsItem]
    return result.data

@app.get("/api/brands/{brand_slug}/search", response_model=Union[schemas.GlobalSearchResult, schemas.SearchNotFoundResponse])
def search_anything(
    brand_slug: str,
    q: str = Query(..., min_length=1),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Tìm kiếm thông minh: Đơn hàng, Mã vận đơn, SĐT hoặc Tên khách hàng.
    """
    db_brand = crud.get_brand_by_slug(db, slug=brand_slug, owner_id=current_user.id)
    if not db_brand:
        raise HTTPException(status_code=404, detail="Không tìm thấy Brand.")
    
    result = search_service.search_entities(db, db_brand.id, q)
    if not result:
        # Trả về 200 kèm status not_found để frontend xử lý êm ái hơn là báo lỗi 404
        return {"status": "not_found", "message": "Không tìm thấy kết quả phù hợp."}
    
    return result

@app.get("/api/brands/{brand_slug}/search-suggestions", response_model=List[schemas.SearchSuggestionItem])
def get_search_suggestions(
    brand_slug: str,
    q: str = Query(..., min_length=2),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Gợi ý nhanh khi người dùng nhập vào ô tìm kiếm.
    """
    db_brand = crud.get_brand_by_slug(db, slug=brand_slug, owner_id=current_user.id)
    if not db_brand:
        raise HTTPException(status_code=404, detail="Không tìm thấy Brand.")
    
    return search_service.suggest_entities(db, db_brand.id, q)

@app.put("/api/brands/{brand_slug}/customers/{customer_identifier}", response_model=schemas.CustomerDetailResponse)
def update_customer_api(
    brand_slug: str,
    customer_identifier: str,
    update_data: schemas.CustomerUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Cập nhật thông tin khách hàng (SĐT, Email, Address, Notes).
    """
    db_brand = crud.get_brand_by_slug(db, slug=brand_slug, owner_id=current_user.id)
    if not db_brand:
        raise HTTPException(status_code=404, detail="Không tìm thấy Brand.")
        
    updated_customer = crud.customer.update_customer_info(db, db_brand.id, customer_identifier, update_data)
    
    if not updated_customer:
        raise HTTPException(status_code=404, detail="Không tìm thấy Khách hàng.")
        
    return search_service.get_customer_profile(db, db_brand.id, updated_customer)